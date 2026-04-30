from datetime import datetime
from models.result_models import EquipmentResult


def generate_excel(result: EquipmentResult, filepath: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl 패키지가 필요합니다: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "설계계산서"

    hdr_fill  = PatternFill("solid", fgColor="1565C0")
    hdr_font  = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
    lbl_fill  = PatternFill("solid", fgColor="E3F2FD")
    lbl_font  = Font(bold=True, name="맑은 고딕", size=9)
    val_font  = Font(name="맑은 고딕", size=9)
    title_font= Font(bold=True, name="맑은 고딕", size=14)
    thin      = Side(style="thin", color="BDBDBD")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")
    right_a   = Alignment(horizontal="right",  vertical="center")

    def set_header(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, center, border

    def set_label(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill, c.font, c.alignment, c.border = lbl_fill, lbl_font, center, border

    def set_val(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font, c.alignment, c.border = val_font, right_a, border

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A1"] = f"사료플랜트 기계 설계 계산서  [{result.equipment_type}]   {ts}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:E1")

    row = 3
    # 모터
    set_header(row, 1, "항목"); set_header(row, 2, "값"); set_header(row, 3, "단위")
    ws.cell(row=row, column=1).value = "▶ 모터 선정 결과"
    ws.cell(row=row, column=1).font = Font(bold=True, color="1565C0", name="맑은 고딕", size=10)
    ws.merge_cells(f"A{row}:C{row}")
    row += 1

    m = result.motor
    for lbl, val, unit in [
        ("필요 동력",      f"{m.required_power_kW:.3f}", "kW"),
        ("선정 모터 용량", f"{m.selected_motor_kW}",    "kW"),
        ("모터 모델",      m.motor_model,                ""),
        ("IEC 프레임",     m.iec_frame,                  ""),
        ("정격 회전수",    f"{m.rated_rpm}",             "rpm"),
        ("정격 전류(400V)",f"{m.rated_current_A}",       "A"),
        ("정격 토크",      f"{m.rated_torque_Nm}",       "N·m"),
        ("효율",           f"{m.efficiency_pct}",        "%"),
    ]:
        set_label(row, 1, lbl); set_val(row, 2, val); set_val(row, 3, unit)
        row += 1

    row += 1
    ws.cell(row=row, column=1).value = "▶ 베어링 선정 결과 (ISO 281)"
    ws.cell(row=row, column=1).font = Font(bold=True, color="1565C0", name="맑은 고딕", size=10)
    ws.merge_cells(f"A{row}:E{row}")
    row += 1
    set_header(row, 1, "항목"); set_header(row, 2, "구동측"); set_header(row, 3, "피동측")
    row += 1

    bd, bdn = result.bearing_drive, result.bearing_driven
    for lbl, a1, a2 in [
        ("베어링 번호",           bd.bearing_number,                  bdn.bearing_number),
        ("제조사",                bd.manufacturer,                    bdn.manufacturer),
        ("내경 (mm)",             f"{bd.bore_mm:.0f}",                f"{bdn.bore_mm:.0f}"),
        ("외경 (mm)",             f"{bd.outer_dia_mm:.0f}",           f"{bdn.outer_dia_mm:.0f}"),
        ("등가하중 P (N)",        f"{bd.equivalent_load_P_N:,.0f}",   f"{bdn.equivalent_load_P_N:,.0f}"),
        ("기본동정격하중 C (N)",  f"{bd.basic_load_rating_C_N:,.0f}", f"{bdn.basic_load_rating_C_N:,.0f}"),
        ("L10 수명 (hr)",         f"{bd.L10_hr:,.0f}",                f"{bdn.L10_hr:,.0f}"),
    ]:
        set_label(row, 1, lbl); set_val(row, 2, a1); set_val(row, 3, a2)
        row += 1

    row += 1
    ws.cell(row=row, column=1).value = "▶ 샤프트 설계 (ASME)"
    ws.cell(row=row, column=1).font = Font(bold=True, color="1565C0", name="맑은 고딕", size=10)
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    s = result.shaft
    for lbl, val, unit in [
        ("재질",          s.material,                     ""),
        ("계산 직경",     f"{s.required_diameter_mm:.2f}", "mm"),
        ("선정 직경(KS)", f"{s.selected_diameter_mm:.0f}", "mm"),
        ("Von Mises 응력",f"{s.von_mises_stress_MPa:.2f}", "MPa"),
        ("허용 응력",     f"{s.allowable_stress_MPa:.2f}", "MPa"),
        ("실제 안전계수", f"{s.safety_factor_actual:.2f}", ""),
    ]:
        set_label(row, 1, lbl); set_val(row, 2, val); set_val(row, 3, unit)
        row += 1

    row += 1
    ws.cell(row=row, column=1).value = "▶ V벨트 선정 (KS B 1400)"
    ws.cell(row=row, column=1).font = Font(bold=True, color="1565C0", name="맑은 고딕", size=10)
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    vb = result.vbelt
    for lbl, val, unit in [
        ("단면",           vb.section,                       ""),
        ("호칭",           vb.belt_length_designation,       ""),
        ("벨트 길이",      f"{vb.belt_length_mm:.0f}",       "mm"),
        ("벨트 수량",      f"{vb.number_of_belts}",          "개"),
        ("구동 풀리 직경", f"{vb.drive_pulley_dia_mm:.0f}",  "mm"),
        ("피동 풀리 직경", f"{vb.driven_pulley_dia_mm:.0f}", "mm"),
        ("접촉각",         f"{vb.contact_angle_deg:.1f}",    "°"),
    ]:
        set_label(row, 1, lbl); set_val(row, 2, val); set_val(row, 3, unit)
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14

    wb.save(filepath)
